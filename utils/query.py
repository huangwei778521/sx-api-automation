import os
from utils.mysql import Mysql
from config.server_info import server_info
from utils.ssh import SSH
from openpyxl import load_workbook


# 查询数据库tidb密码 & host_url
def query_server_info():
    host_info = server_info.get("host_info")
    host = host_info["host"]
    user = host_info["user"]
    password = host_info["password"]
    ssh = SSH(host, user, password)
    ssh.conn()
    get_tidb_password = "kubectl get secrets password-secrets -o yaml | grep tidb_root_password | awk '{print $2}' | " \
                        "base64 -d "
    get_host_url = "kubectl get ingress.|grep business|awk '{print $3}'"
    password = ssh.run_cmd(get_tidb_password)
    host_url = ssh.run_cmd(get_host_url)
    ssh.close()
    return password[0], host_url[0]


# 查询设备信息
def query_device_info():
    sql = 'SELECT device_name from info_device'
    with Mysql() as mysql:
        query_result = mysql.select_all(sql)
    list_name = []
    for query_name in query_result:
        list_name.append(query_name[0].decode())
    return list_name


# 根据设备名称查询设备id
def query_device_id(device_name):
    sql = f"SELECT device_id  from info_device where device_name ='{device_name}' and device_status ='1'"
    with Mysql() as mysql:
        query_result = mysql.select_one(sql)
    return query_result[0]


# 获取目录下子目录名称
def query_dirs(root_dir):
    for parent, dir_names, filenames in os.walk(root_dir):
        return dir_names


# 获取目录下子文件名称
def query_files(root_dir):
    for parent, dir_names, filenames in os.walk(root_dir):
        print(filenames)


# 读取test.conf文件并输出文件名
def read_config(file_path):
    f = open(file_path, 'r', encoding='UTF-8')
    list_videos = []
    for i in f:
        if i[0:7] == '<Stream':
            list_videos.append(i.split('<Stream ')[1].split('>\n')[0])
        else:
            pass
    return list_videos


# 查询用户id与租户id
def query_root_user_info():
    sql = "SELECT user_id ,tenant_id  from user where type='TENANT_ADMIN'"
    with Mysql() as mysql:
        query_result = mysql.select_one(sql)
    return [query_result[0], query_result[1].decode()]


# 通过task_name检索对应的传参json
def query_payload(task_name):
    file_path = '/Users/liuyang5/PycharmProjects/sx-api-automation/config/business/autotest.xlsx'
    wb = load_workbook(file_path)
    sheet = wb.active
    col_A = sheet['A']
    for cell in col_A:
        if cell.value == f'{task_name}':
            value_right = sheet.cell(row=cell.row, column=cell.column + 1).value
            return eval(value_right)
        else:
            pass


# 通过设备id检索task_id
def query_task_id(device_id):
    sql = f"SELECT task_id from device_task_config where device_id ={device_id} "
    with Mysql() as mysql:
        query_result = mysql.select_one(sql)
    return query_result[0]


# 查询总园区地图id
def query_entire_zone_id():
    sql = "SELECT id  from map_zone mz where parent_id =0 "
    with Mysql() as mysql:
        query_result = mysql.select_one(sql)
    return query_result[0]


